import pandas as pd
from pathlib import Path
import openpyxl

# Use Path to handle unicode paths better
base = Path("config")
target = base / "动物界-脊索动物门-2025-10626.xlsx"

print(f"Target exists: {target.exists()}")

try:
    wb = openpyxl.load_workbook(target, read_only=True)
    print("Sheet names:", wb.sheetnames)
    
    if wb.sheetnames:
        df = pd.read_excel(target, sheet_name=wb.sheetnames[0])
        print("Columns:", list(df.columns))
except Exception as e:
    print(f"Error: {e}")
